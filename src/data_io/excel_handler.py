from openpyxl import Workbook, load_workbook

def read_excel(file_path):
    """Read an Excel file and return a dictionary of {sheet_name: data}"""
    workbook = load_workbook(filename=file_path)
    
    # Create a dictionary to store data from each sheet
    all_sheets_data = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []
        
        for row in sheet.iter_rows(values_only=True):
            # Convert None values to empty strings
            processed_row = ['' if cell is None else str(cell) for cell in row]
            sheet_data.append(processed_row)
            
        all_sheets_data[sheet_name] = sheet_data
    
    return all_sheets_data

def write_excel(file_path, data):
    """
    Write data to an Excel file
    
    Args:
        file_path: Path to save the file
        data: Either a list of lists (single sheet) or a dict of {sheet_name: sheet_data}
    """
    workbook = Workbook()
    
    # Remove the default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    if isinstance(data, dict):
        # Handle multiple sheets
        for sheet_name, sheet_data in data.items():
            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)
    else:
        # Handle single sheet
        sheet = workbook.create_sheet(title="Sheet1")
        for row in data:
            sheet.append(row)
    
    workbook.save(file_path)